import os
import zipfile
from pypdf import PdfReader
from zipfile import ZipFile
from openpyxl.reader.excel import load_workbook
import csv

from utils.constants import (
    RESOURCES_DIR,
    ARCHIVE_PATH,
    PDF_FILE,
    CSV_FILE,
    XLSX_FILE,
    LIST_OF_FILES,
)


def test_archive_exists(create_archive):
    assert os.path.exists(ARCHIVE_PATH)


def test_check_files_in_archive(create_archive):
    archive_zip_list = ZipFile(ARCHIVE_PATH).namelist()
    for zipped_file in archive_zip_list:
        assert zipped_file in LIST_OF_FILES


def test_pdf_contents(create_archive):
    with zipfile.ZipFile(ARCHIVE_PATH, "r") as archive:
        pdf_file = next((f for f in archive.namelist() if f.endswith(".pdf")), None)
        assert pdf_file is not None, "PDF file is missing in the archive."
        with archive.open(pdf_file) as file_in_archive:
            archived_pdf = PdfReader(file_in_archive)
            expected_text = archived_pdf.pages[0].extract_text()
            original_pdf = PdfReader(os.path.join(RESOURCES_DIR, PDF_FILE))
            original_text = original_pdf.pages[0].extract_text()
            assert expected_text == original_text, "PDF content does not match."


def test_xlsx_contents(create_archive):
    with zipfile.ZipFile(ARCHIVE_PATH, "r") as archive:
        xlsx_file = next((f for f in archive.namelist() if f.endswith(".xlsx")), None)
        assert xlsx_file is not None, "XLSX file is missing in the archive."
        with archive.open(xlsx_file) as file_in_archive:
            wb = load_workbook(file_in_archive)
            sheet = wb.active
            archived_value = sheet["B4"].value
            wb_orig = load_workbook(os.path.join(RESOURCES_DIR, XLSX_FILE))
            sheet_orig = wb_orig.active
            original_value = sheet_orig["B4"].value
            assert archived_value == original_value, "XLSX cell content does not match."


def test_csv_contents(create_archive):
    with zipfile.ZipFile(ARCHIVE_PATH, "r") as archive:
        assert any(
            f.endswith(".csv") for f in archive.namelist()
        ), "CSV file is missing in the archive."
        with archive.open(CSV_FILE) as file_in_archive, open(
            os.path.join(RESOURCES_DIR, CSV_FILE), newline=""
        ) as original_file:
            archive_csv_reader = csv.reader(
                file_in_archive.read().decode("utf-8").splitlines()
            )
            original_csv_reader = csv.reader(original_file)
            assert list(archive_csv_reader) == list(
                original_csv_reader
            ), "CSV content does not match."
